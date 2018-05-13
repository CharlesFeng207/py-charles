# coding=utf-8
from os import path
from itertools import groupby
from datetime import datetime
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
import shutil

def sort_and_groupby(obj_lst, func):
    obj_lst.sort(key=func)
    lstg = groupby(obj_lst,key = func)
    return {k:list(g) for k,g in lstg}

def combine_key_to_list(obj_lst, func, number_id = None):

    temp_obj_dic = {} # { id_value : { every_key : [every_values] } }

    for obj in obj_lst:
        
        combine_id_value = func(obj)

        # if this is a new value, add it
        if combine_id_value not in temp_obj_dic:
            temp_obj_dic[combine_id_value] = number_id is None and {} or {number_id:0}
            
        key_list_obj = temp_obj_dic[combine_id_value]

        # add number
        if number_id is not None:
            key_list_obj[number_id] += 1

        # handle other keys for this value
        for k in obj:
            if k not in key_list_obj:
                key_list_obj[k] = []

            if obj[k] not in key_list_obj[k]:
                key_list_obj[k].append(obj[k])
    
    return map(lambda x:{k : x[k] for k in x}, temp_obj_dic.values())

def attach_number_col(obj_lst, num):
    for i, obj in enumerate(obj_lst):
        obj[num] = i + 1

def to_str(obj):

    if obj == None:
        return ""

    if type(obj) is str:
        return obj

    if type(obj) is unicode:
        return obj.encode('UTF-8')

    if type(obj) is datetime:
        return obj.strftime("%Y-%m-%d")

    if type(obj) is list:
        return ','.join(map(to_str, obj))

    return str(obj)

def make_table(template_path, output_path, id_row, writing_row, obj_lst):
    
    shutil.copy(template_path, output_path)

    workbook_target = load_workbook(output_path)
    sheet_target = workbook_target.active

    column_letters = map(lambda i: get_column_letter(i), range(1, sheet_target.max_column + 1))
    column_letters_id_dic = { k:sheet_target["{}{}".format(k, id_row)].value for k in column_letters}

    for obj in obj_lst:
        for letter in column_letters_id_dic.keys():
            cell_id = column_letters_id_dic[letter]
            if cell_id in obj:
                sheet_target["{}{}".format(letter, writing_row)].value = to_str(obj[cell_id])
           
        writing_row += 1
    
    workbook_target.save(output_path)

    print "{} created!".format(output_path)

if __name__ == '__main__':

    friends = [191, 158, 159, 165, 170, 177, 181, 182, 190]
    print sort_and_groupby(friends, lambda x: x > 160 and 'tall' or 'short')

    print '------'

    d1={'name':'zhangsan','age':20,'country':'China'}
    d2={'name':'wangwu','age':19,'country':'USA'}
    d3={'name':'lisi','age':22,'country':'JP'}
    d4={'name':'zhaoliu','age':20,'country':'USA'}
    d5={'name':'pengqi','age':22,'country':'USA'}
    d6={'name':'lijiu','age':20,'country':'China'}
    lst=[d1,d2,d3,d4,d5,d6]

    print combine_key_to_list(lst, lambda x:x['country'])
