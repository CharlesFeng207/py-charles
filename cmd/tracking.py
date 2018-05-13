# coding:utf8

from datetime import datetime
import json
import shutil
import sys
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
import charlesUtil
from charlesUtil import to_str

def find_tracking_row_objs():

    # indicate from which row number would be known as data
    src_data_start_row = int(global_json_data["src_data_start_row"])

    # indicate which cols in src table would be judged by time_start and time_end
    # if any of cols satisfy would be selected
    time_range_cols = table_json_data["time_range_cols"].split()

    column_letters = map(lambda i: get_column_letter(i),
                         range(1, sheet_src.max_column + 1))
    column_letters_id_dic = {k: src_letter_to_id(k) for k in column_letters}

    objs = []

    for row in range(src_data_start_row, sheet_src.max_row + 1):
        # will select this row if any of column achieve condition
        for col in time_range_cols:
            cell_name = "{}{}".format(col, row)
            cell_value = sheet_src[cell_name].value

            if cell_value == None:
                continue

            # print cell_value, type(cell_value)

            if check_time_range(cell_value):
                cell_obj = {}  # collect all data as a diectionary

                for column_letter in column_letters:  # each column as a key
                    column_id = column_letters_id_dic[column_letter]

                    if column_id == "None":
                        continue

                    cell_obj[column_id] = sheet_src["{}{}".format(column_letter, row)].value
                    pass
                
                objs.append(cell_obj)
                print "row {} selected ({})".format(row, cell_value)
                break  # avoid duplicated selecting

    return objs

def check_time_range(cell_value):
    
    # it's string, covert to datetime, may be multiple data
    if type(cell_value) is unicode:
        t1 = map(lambda x: datetime.strptime(x, "%Y/%m/%d"), cell_value.split())
        t2 = map(lambda x: check_time_range(x), t1)
        t3 = any(t2) # as long as one of data successful
        return t3

    if type(cell_value) is datetime:
        if cell_value >= time_start and cell_value <= time_end:
            return True

    return False

def process_working_table():

    # if not define table name means it's not requred to make this table
    working_table_name = table_json_data["working_table_name"]
   
    if working_table_name == u'':
        print "working_table_name is null, it's not requred to make this table"
        return
    
    # if the combine id of two records is the same, they should be combined to one
    working_combine_id = src_letter_to_id(table_json_data["working_combine_col"])

    # provide a property id to keep record cout after combined
    working_combine_number_id = table_json_data["working_combine_number_id"]

    after_combined = charlesUtil.combine_key_to_list(src_selected_objs, lambda x:x[working_combine_id], working_combine_number_id)
    charlesUtil.attach_number_col(after_combined, 'No.')

    # prepare parameter to make table
    temp_working = table_json_data["temp_working"]
    working_output_path = "{}{}({} to {}).xlsx".format(user_output_folder, working_table_name, to_str(time_start), to_str(time_end))
    temp_working_id_row = int(table_json_data["temp_working_id_row"])
    temp_working_data_start_row = int(table_json_data["temp_working_data_start_row"])

    charlesUtil.make_table(temp_working, working_output_path,temp_working_id_row,temp_working_data_start_row, after_combined)

def process_delay_table():

    delay_table_name = table_json_data["delay_table_name"]

    if delay_table_name == u'':
        print "delay_table_name is null"
        return

    # just select col which isn't filled with data
    delay_filter_id = src_letter_to_id(table_json_data["delay_check_col"])
    after_filtered = filter(lambda x:x[delay_filter_id] == None, src_selected_objs)

    # if the combine id of two records is the same, they should be combined to one
    delay_combine_id = src_letter_to_id(table_json_data["delay_combine_col"])
    
    # provide a property id to keep record cout after combined
    delay_combine_number_id = table_json_data["delay_combine_number_id"]

    after_filtered_combined = charlesUtil.combine_key_to_list(after_filtered, lambda x:x[delay_combine_id], delay_combine_number_id)
    charlesUtil.attach_number_col(after_filtered_combined, 'No.')

    # prepare parameter to make table
    delay_output_path = "{}{}({} to {}).xlsx".format(user_output_folder, delay_table_name, to_str(time_start), to_str(time_end))
    temp_delay_path = table_json_data["temp_delay"]
    temp_delay_id_row = int(table_json_data["temp_delay_id_row"])
    temp_delay_data_start_row = int(table_json_data["temp_delay_data_start_row"])

    charlesUtil.make_table(temp_delay_path,delay_output_path, temp_delay_id_row, temp_delay_data_start_row, after_filtered_combined)

def src_letter_to_id(letter):
    return sheet_src["{}{}".format(letter, src_id_row)].value

if len(sys.argv) == 4:
    
    print "sys.argv:", sys.argv

    # get user parameters
    global_json_path = sys.argv[1]
    table_json_path = sys.argv[2]
    user_output_folder = sys.argv[3]

    # init json config
    with open(global_json_path) as global_json_file:
        print "global json config loaded: ", global_json_path

        # hanle backsplash
        global_json_str = global_json_file.read().replace("\\", "\\\\")
        print global_json_str, type(global_json_str)

        global_json_data = json.loads(global_json_str)

        with open(table_json_path) as table_json_file:
            print "table json config loaded: ", table_json_path

            table_json_str = table_json_file.read().replace("\\", "\\\\")
            print table_json_str, type(table_json_str)

            table_json_data = json.loads(table_json_str)

            # read time info
            time_start_str = global_json_data["time_start"]
            time_end_str = global_json_data["time_end"]
            time_start = datetime.strptime(time_start_str, "%Y/%m/%d")
            time_end = datetime.strptime(time_end_str, "%Y/%m/%d")

            # load source excel
            path_src = global_json_data["src"]
            print "loading {}".format(path_src)

            workbook_src = load_workbook(path_src)
            sheet_src = workbook_src[global_json_data["src_sheet"]]
            src_id_row = global_json_data["src_id_row"]

            print "load complete! wait for processing ..."

            # find tracking rows by time info and save as dictionary
            src_selected_objs = find_tracking_row_objs()

            print "from {} to {}, there are {} rows selected!".format(time_start_str, time_end_str, len(src_selected_objs))

            # make tables by 'src_selected_objs'

            # make working table
            process_working_table()

            # make delay table
            process_delay_table()
       

else:
    print "please input three parameters! (global json path, table json path, out put folder path)"
