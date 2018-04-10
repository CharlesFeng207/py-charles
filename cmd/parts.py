# coding:utf8

import json
import shutil
import time
import traceback
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from copy import deepcopy

class PartChanging:
    
    op_init = 'init'
    op_override = 'override' 
    op_update = 'update'
    op_delete = 'delete' 

    def __init__(self, new_part_id, old_part_id, new_part_data, old_part_data, row_number):
        self.new_part_id = new_part_id
        self.old_part_id = old_part_id
        self.old_part_data = old_part_data
        self.new_part_data = new_part_data
        self.row_number = row_number

    def __str__(self):
         return '< PartChanging id:{} {} -> {} {} row:{} >'.format(
             self.old_part_id, is_valid_id(self.old_part_id), self.new_part_id, is_valid_id(self.new_part_id), self.row_number)
    
    def do(self, parts_wrapper_objs):

        print self

        # when a part first time to appear
        PartChanging.init_before_operation(parts_wrapper_objs, self.old_part_id, self.old_part_data, self.row_number)

        PartChanging.init_before_operation(parts_wrapper_objs, self.new_part_id, self.new_part_data, self.row_number)

        # if new part id is invalid but the old is valid will clear all usage for old part id
        if (not is_valid_id(self.new_part_id)) and is_valid_id(self.old_part_id): # delete
            empty_data = PartDataRecord()
            empty_data.record_info = '{} {} (empty) by row {}'.format(PartChanging.op_delete, self.old_part_id, self.row_number)
            print empty_data.record_info
            parts_wrapper_objs[self.old_part_id].append_part_data_record(empty_data)

        # if new part is just old part, will override old part data
        elif is_valid_id(self.new_part_id) and is_valid_id(self.old_part_id) and self.new_part_id == self.old_part_id: # override
            modified_data = deepcopy(self.new_part_data)
            modified_data.record_info = '{} {} ({} -> {}) by row {}'.format(
                PartChanging.op_override, self.old_part_id, 
                self.old_part_data.data_summary, self.new_part_data.data_summary, self.row_number)

            print modified_data.record_info

            parts_wrapper_objs[self.old_part_id].append_part_data_record(modified_data)

        # if old part updated to new part, will clear usage which the new part using, 
        # after that build a relationship between the new and old
        elif is_valid_id(self.new_part_id) and is_valid_id(self.old_part_id) and self.new_part_id != self.old_part_id:

            modified_data = deepcopy(self.old_part_data)

            for k in self.new_part_data.car_usage:
                if k in modified_data.car_usage and self.new_part_data.car_usage[k] > 0:
                    modified_data.car_usage[k] = 0

            old_part_state = parts_wrapper_objs[self.old_part_id]
            new_part_state = parts_wrapper_objs[self.new_part_id]

            modified_data.record_info = '{} {}({} -> {}) by row {}, new part id: {}'.format(
                PartChanging.op_update, self.old_part_id, 
                self.old_part_data.data_summary, modified_data.data_summary, 
                self.row_number, self.new_part_id)

            print modified_data.record_info

            old_part_state.append_part_data_record(modified_data)

            # build relationship
            if new_part_state not in old_part_state.next_part:
                old_part_state.next_part.append(new_part_state)

            new_part_state.pre_part = old_part_state
        else:
            print 'unhandle situaltion!'
            
        print '\n'
           
    @staticmethod
    def init_before_operation(parts_wrapper_objs, part_id, part_data, row_number):
        if is_valid_id(part_id) and part_id not in parts_wrapper_objs:
            init_data = deepcopy(part_data)

            init_data.record_info = '{} {}({}) by row {}'.format(
                PartChanging.op_init, part_id, init_data.data_summary, row_number)
            
            print init_data.record_info

            t = PartWrapper(part_id)
            t.append_part_data_record(init_data)
            parts_wrapper_objs[part_id] = t

class PartWrapper:
    
    def __init__(self, part_id):
        self.pre_part = None
        self.next_part = []
        self.part_id = part_id
        self.part_data_records = []

    def __str__(self):
        pre_str = self.pre_part.part_id if self.pre_part != None else 'None'
        root_str = self.root_wrapper.part_id if self.root_wrapper != None else 'None'

        next_str = str(map(lambda x: x.part_id, self.next_part))
        return "< PartWrapper id:{} pre:{} root:{} next:{} avalible:{}>".format(
            str(self.part_id), pre_str, root_str, next_str, self.is_avalible)

    def append_part_data_record(self, part_data_record):
        self.part_data_records.append(part_data_record)

    @property
    def detail_str(self):
        arr = [str(item) for item in self.part_data_records]
        arr.insert(0, str(self)) # add self
        arr_str = '\n'.join(arr)
        return arr_str

    @property
    def is_root_wrapper(self):
        return self.pre_part == None

    @property
    def is_final_wrapper(self):
        return len(self.next_part) == 0
    
    @property
    def root_wrapper(self):
        t = self
        while not t.is_root_wrapper:
            t = t.pre_part
        return t

    @property
    def initial_part_data(self):
        if len(self.part_data_records) > 0:
            return self.part_data_records[0]
        return None
    
    @property
    def newest_part_data(self):
        if len(self.part_data_records) > 0:
            return self.part_data_records[-1]
        return None

    @property
    def is_avalible(self):
        t = self.newest_part_data
        return t.is_avalible if t != None else False

class PartDataRecord:
    
    def __init__(self):
        self.car_usage = {}
        self.record_info = None

    def __str__(self):
        return "< {} >".format(self.record_info)
    
    def set_usage(self, car_usage):
        self.car_usage = car_usage
        pass

    @property
    def is_avalible(self):
        t = map(lambda x: x > 0, self.car_usage.values())
        return any(t)

    @property
    def car_usage_str(self):
        arr = [str(self.car_usage[column_letters_id_dic[col]]) for col in new_part_usage_cols if column_letters_id_dic[col] in self.car_usage]
        arr_str = ''.join(arr)
        return arr_str
    
    @property
    def data_summary(self):
        return self.car_usage_str

def src_letter_to_id(letter):
    return sheet_src["{}{}".format(letter, src_id_row)].value

def load_cell_part_id(sheet, col, row):
    cell_value = sheet["{}{}".format(col, row)].value

    if type(cell_value) is unicode:
        cell_value = u"".join(cell_value.split()) # delete nbsp
        return cell_value.encode('utf-8')

    if type(cell_value) is str:
        return cell_value

    if cell_value == None:
        return None
        
    return str(cell_value)

def load_cell_car_usage(sheet, col, row):
    
    cell_value = sheet["{}{}".format(col, row)].value

    if cell_value == None:
        return 0

    # check if it is a number
    if not any(map(lambda x:type(cell_value) is x, [float, int, long])):
        print "error: {}{} {}({}) is not number!!".format(col, row, cell_value, type(cell_value))
        raw_input()

    return cell_value

def print_parts_wrapper_objs(parts_wrapper_objs):

    print "< print_parts_wrapper_objs >\n"

    for item in parts_wrapper_objs.values():
        print item.detail_str, "\n"
    
    print "\n< print_parts_wrapper_objs end >\n"

def is_valid_id(part_id):
    if type(part_id) is str:
        t1 = part_id.split('-')
        if len(t1) > 2:
            t2 = map(lambda x:x == '0' or x == '', t1)
            t3 = not all(t2)
            return t3
    
    return False

def load_parts_wookbook(workbook_path):
    print 'loading... ', workbook_path
    since = time.time()
    workbook_loaded = load_workbook(workbook_path, data_only=True)
    print "load complete! time cost: ", time.time() - since

    return workbook_loaded

def process_parts_initial_table():
    
    parts_table_path = u'D:\\Repositories\\py_charles\\cmd\\C490 MCA TT BoM Validation_20180319.xlsx'
    workbook_parts_target = load_parts_wookbook(parts_table_path)
    sheet_name = 'TT BoM '
    sheet_parts_target = workbook_parts_target[sheet_name]

    targert_parts_col = 'AP'
    row_start = 10
    row_end = 1562

    taret_parts_id = [(row, load_cell_part_id(sheet_parts_target, targert_parts_col, row)) for row in range(row_start, row_end + 1)]
    
    for row, target_id in taret_parts_id:
        if not is_valid_id(target_id):
            print 'error: {} (row {}) is invalid!'.format(target_id, row)
            raw_input()
            continue
        
        if target_id not in parts_wrapper_objs:
            print '{} (row {}) not found in src wookbook'.format(target_id, row)
            # raw_input()
            continue

        print 'the root of target {} (row {}) is {}\n'.format(target_id, row, parts_wrapper_objs[target_id].root_wrapper)
        
src_path = u'D:\\Repositories\\py_charles\\cmd\\1234.xlsx'
workbook_src = load_parts_wookbook(src_path)
sheet_src = workbook_src['Change Log']

src_id_row = 5
src_data_start_row = 7

new_part_id_col = 'M'
old_part_id_col = 'AR'
new_part_usage_cols = ['O','P','Q','R','S','T','U','V']
old_part_usage_cols = ['AT','AU','AV','AW','AX','AY','AZ','BA']

column_letters = map(lambda i: get_column_letter(i), range(1, sheet_src.max_column + 1))
column_letters_id_dic = {k: src_letter_to_id(k) for k in column_letters}

changelist = [] # collect change logs
for row in range(src_data_start_row, sheet_src.max_row + 1):
    
    new_part_id = load_cell_part_id(sheet_src, new_part_id_col, row)
    new_part_usage = {}
    for col in new_part_usage_cols:
        cell_value = load_cell_car_usage(sheet_src, col, row)
        new_part_usage[column_letters_id_dic[col]] = cell_value

    new_part_data = PartDataRecord()
    new_part_data.set_usage(new_part_usage)

    old_part_id = load_cell_part_id(sheet_src, old_part_id_col, row)
    old_part_usage = {}
    for col in old_part_usage_cols:
        cell_value = load_cell_car_usage(sheet_src, col, row)
        old_part_usage[column_letters_id_dic[col]] = cell_value
    
    old_part_data = PartDataRecord()
    old_part_data.set_usage(old_part_usage)

    part_change = PartChanging(new_part_id, old_part_id, new_part_data, old_part_data, row)
    changelist.append(part_change)

try:
    parts_wrapper_objs = {}
    
    print len(changelist), ' change logs found! \n'
    
    for item in changelist:
        item.do(parts_wrapper_objs)
    
    print 'build parts object complete! \n \n'

    print_parts_wrapper_objs(parts_wrapper_objs)

    process_parts_initial_table()

except Exception as err:
    print err
    print traceback.format_exc()

raw_input()